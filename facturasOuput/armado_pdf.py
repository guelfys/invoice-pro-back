from backend.log import escribir_log, obtener_timestamp
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import win32com.client as win32
import os
import shutil
import re
import json
import base64
import datetime as _dt
import tempfile
from io import BytesIO
from facturasOuput.qr_rescue import asegurar_qr_en_factura
import pythoncom

try:
    import qrcode  
except Exception:
    qrcode = None

_QR_BASE_URL = "https://www.arca.gob.ar/fe/qr/"

def _solo_digitos(v):
    if v is None:
        return ""
    return re.sub(r"\D+", "", str(v))

def _to_int_safe(v, default=None):
    s = _solo_digitos(v)
    if s == "":
        return default
    try:
        return int(s)
    except Exception:
        return default

def _to_float_safe(v, default=0.0):
    if v is None:
        return default
    if isinstance(v, str):
        s = v.strip()
        if s == "":
            return default
        # soporta "1.234,56" y "1234.56"
        s = s.replace(".", "").replace(",", ".") if ("," in s and "." in s) else s.replace(",", ".")
        try:
            return float(s)
        except Exception:
            return default
    try:
        return float(v)
    except Exception:
        return default

def _set_num(ws, cell_ref, value, fmt="#,##0.00"):

    try:
        if value is None:
            v = 0.0
        elif isinstance(value, str):
            s = value.strip()
            if s == "":
                v = 0.0
            else:
                # soporta "1.234,56" o "1234.56" o "1234,56"
                if "," in s and "." in s:
                    s = s.replace(".", "").replace(",", ".")
                else:
                    s = s.replace(",", ".")
                v = float(s)
        else:
            v = float(value)
    except Exception:
        v = 0.0

    c = ws[cell_ref]
    c.value = v
    c.number_format = fmt



def _normalize_key(value):
    if value is None:
        return ""
    return (
        str(value)
        .strip()
        .lower()
        .replace("\n", " ")
        .replace("\r", " ")
    )

def _row_lookup(row):
    try:
        idx = row.index
    except Exception:
        idx = []
    return {_normalize_key(col): col for col in idx if col is not None}

def _get_by_index(row, idx, default=None):
    try:
        if idx is None:
            return default
        value = row.iloc[idx]
        if pd.isna(value):
            return default
        return value
    except Exception:
        return default

def _get_row_value(row, *possible_names, default=None):
    lookup = _row_lookup(row)
    for name in possible_names:
        real_col = lookup.get(_normalize_key(name))
        if real_col is not None:
            try:
                value = row.get(real_col)
            except Exception:
                value = None
            if value is not None and not pd.isna(value):
                if not (isinstance(value, str) and value.strip() == ""):
                    return value
    return default

def _get_text(row, *possible_names, default=""):
    value = _get_row_value(row, *possible_names, default=None)
    if value is None:
        return default
    try:
        if pd.isna(value):
            return default
    except Exception:
        pass
    return str(value).strip()

def _safe_text(value, default=""):
    if value is None:
        return default
    try:
        if pd.isna(value):
            return default
    except Exception:
        pass
    s = str(value).strip()
    return s if s else default

def _coerce_date_for_excel(value):
    if value is None:
        return None

    try:
        if pd.isna(value):
            return None
    except Exception:
        pass

    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()

    if isinstance(value, _dt.datetime):
        return value

    if isinstance(value, _dt.date):
        return _dt.datetime.combine(value, _dt.time())

    if isinstance(value, (int, float)):
        try:
            fv = float(value)
            if 20000101 <= fv <= 29991231 and float(value).is_integer():
                s = str(int(fv))
                return _dt.datetime(int(s[0:4]), int(s[4:6]), int(s[6:8]))
            if 30000 <= fv <= 80000:
                return _dt.datetime(1899, 12, 30) + _dt.timedelta(days=fv)
        except Exception:
            return None

    s = str(value).strip()
    if not s:
        return None

    if re.fullmatch(r"\d{8}", s):
        try:
            return _dt.datetime(int(s[0:4]), int(s[4:6]), int(s[6:8]))
        except Exception:
            return None

    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return _dt.datetime.strptime(s, fmt)
        except Exception:
            pass

    try:
        return _dt.datetime.fromisoformat(s)
    except Exception:
        return None

def _set_excel_date(ws, cell_ref, value):
    ws[cell_ref] = _coerce_date_for_excel(value)

def _sheet_name_input(tipo_nota, tipo_factura):
    tipo_nota_norm = _safe_text(tipo_nota).lower()
    tipo_factura_norm = _safe_text(tipo_factura).upper()

    if tipo_nota_norm == "factura":
        return f"Factura {tipo_factura_norm}"
    if tipo_nota_norm == "credito":
        return f"Nota Credito {tipo_factura_norm}"
    if tipo_nota_norm == "debito":
        return f"Nota Debito {tipo_factura_norm}"
    return f"Factura {tipo_factura_norm}"

def _legacy_field_positions(tipo_nota, tipo_factura):
    tipo_nota_norm = _safe_text(tipo_nota).lower()
    tipo_factura_norm = _safe_text(tipo_factura).upper()

    if tipo_nota_norm == "factura":
        return {
            "fecha_emision": 1,
            "periodo_desde": 2,
            "periodo_hasta": 3,
            "fecha_vencimiento_pago": 4,
            "condicion_iva": 5,
            "tipo_doc": 6,
            "cliente": 7,
            "documento": 8,
            "domicilio": 9,
            "cantidad": 10,
            "descripcion": 11,
            "precio_unitario": 12,
            "importe_total": 13,
            "condicion_venta": 14,
            "unidad_medida": 15,
            "bonificacion_porcentaje": 16,
            "importe_bonificacion": 17,
            "importe_op_ex": 18,
            "importe_iva": 19,
            "importe_tributos": 20,
            "observaciones": 21,
            "concepto": 22,
            "codigo_condicion_iva": 28,
        }

    if tipo_factura_norm in ("A", "B"):
        return {
            "fecha_emision": 4,
            "periodo_desde": 5,
            "periodo_hasta": 6,
            "fecha_vencimiento_pago": 7,
            "condicion_iva": 3,
            "tipo_doc": 8,
            "cliente": 9,
            "documento": 10,
            "domicilio": 11,
            "unidad_mtx": 12,
            "codigo_mtx": 13,
            "descripcion": 14,
            "cantidad": 15,
            "precio_unitario": 16,
            "importe_total": 17,
            "codigo_condicion_iva": 18,
            "importe_iva": 19,
            "concepto": 20,
            "motivo_nota": 21,
            "importe_otros_tributos": 22,
        }

    return {
        "fecha_emision": 2,
        "periodo_desde": 3,
        "periodo_hasta": 4,
        "fecha_vencimiento_pago": 5,
        "tipo_doc": 6,
        "cliente": 7,
        "documento": 8,
        "domicilio": 9,
        "importe_total": 10,
        "concepto": 11,
        "motivo_nota": 12,
    }

def _resolve_comprobante_fields(comprobante, tipo_nota, tipo_factura):
    legacy = _legacy_field_positions(tipo_nota, tipo_factura)

    def pick(field, *names):
        value = _get_row_value(comprobante, *names, default=None)
        if value is not None:
            return value
        return _get_by_index(comprobante, legacy.get(field), default=None)

    return {
        "fecha_emision": pick("fecha_emision", "Fecha", "Fecha Emisión", "Fecha Emision"),
        "periodo_desde": pick("periodo_desde", "Periodo Desde", "Fecha servicio desde"),
        "periodo_hasta": pick("periodo_hasta", "Periodo Hasta", "Fecha servicio hasta"),
        "fecha_vencimiento_pago": pick("fecha_vencimiento_pago", "Fecha vencimiento pago"),
        "condicion_iva": pick("condicion_iva", "Condicion frente al IVA", "Condición frente al IVA"),
        "tipo_doc": pick("tipo_doc", "Tipo Doc"),
        "cliente": pick("cliente", "Cliente", "Razón Social", "Razon Social"),
        "documento": pick("documento", "Documento"),
        "domicilio": pick("domicilio", "Domicilio"),
        "cantidad": pick("cantidad", "Cantidad", "Cant."),
        "descripcion": pick("descripcion", "Descripcion", "Descripción"),
        "precio_unitario": pick("precio_unitario", "Precio Unitario", "$ Unit."),
        "importe_total": pick("importe_total", "Importe Total", "Total"),
        "condicion_venta": pick("condicion_venta", "Condición Venta", "Condicion Venta"),
        "unidad_medida": pick("unidad_medida", "Unidad Medida"),
        "bonificacion_porcentaje": pick("bonificacion_porcentaje", "% Bonificación", "% Bonificacion"),
        "importe_bonificacion": pick("importe_bonificacion", "Importe Bonificación", "Importe Bonificacion"),
        "importe_op_ex": pick("importe_op_ex", "Importe Op Ex"),
        "importe_iva": pick("importe_iva", "Importe IVA"),
        "importe_tributos": pick("importe_tributos", "Importe Tributos", "Importe Otros Tributos"),
        "observaciones": pick("observaciones", "Observaciones"),
        "concepto": pick("concepto", "Concepto"),
        "motivo_nota": pick("motivo_nota", "Motivo Nota"),
        "codigo_condicion_iva": pick("codigo_condicion_iva", "Codigo Condición IVA", "Codigo Condicion IVA"),
    }

def _codigo_iva_a_texto(codigo):
    codigo = _safe_text(codigo)
    mapa = {
        "3": "0%",
        "4": "10,50%",
        "5": "21%",
        "6": "27%",
        "8": "5%",
        "9": "2.50%",
    }
    return mapa.get(codigo, "")

def _extract_array_items(solicitud):
    try:
        if "arrayItems" not in solicitud.index:
            return []
    except Exception:
        return []

    value = solicitud.get("arrayItems")
    if value is None:
        return []
    try:
        if pd.isna(value):
            return []
    except Exception:
        pass

    items = []
    if isinstance(value, list):
        for item_group in value:
            if isinstance(item_group, dict):
                inner = item_group.get("item", [])
                if isinstance(inner, list):
                    items.extend(inner)
                elif isinstance(inner, dict):
                    items.append(inner)
    elif isinstance(value, dict):
        inner = value.get("item", [])
        if isinstance(inner, list):
            items.extend(inner)
        elif isinstance(inner, dict):
            items.append(inner)

    return items

def _build_single_item_from_comprobante(comprobante, tipo_nota, tipo_factura):
    fields = _resolve_comprobante_fields(comprobante, tipo_nota, tipo_factura)

    cantidad = _to_float_safe(fields.get("cantidad"), default=0.0)
    descripcion = _safe_text(fields.get("descripcion"))
    precio_unitario = _to_float_safe(fields.get("precio_unitario"), default=0.0)
    importe_total = _to_float_safe(fields.get("importe_total"), default=0.0)
    importe_bonificacion = _to_float_safe(fields.get("importe_bonificacion"), default=0.0)
    importe_iva = _to_float_safe(fields.get("importe_iva"), default=0.0)
    bonificacion_porcentaje = _to_float_safe(fields.get("bonificacion_porcentaje"), default=0.0)
    unidad_medida = _safe_text(fields.get("unidad_medida"))
    codigo_condicion_iva = _safe_text(fields.get("codigo_condicion_iva"))

    if cantidad == 0 and importe_total:
        cantidad = 1.0
    if precio_unitario == 0 and cantidad:
        precio_unitario = importe_total / cantidad if cantidad else importe_total

    if not descripcion:
        descripcion = (
            _safe_text(fields.get("motivo_nota"))
            or _safe_text(fields.get("observaciones"))
            or f"{_safe_text(tipo_nota, 'Comprobante')} {_safe_text(tipo_factura)}"
        )

    return {
        "codigo": 1,
        "descripcion": descripcion,
        "cantidad": cantidad if cantidad else 1.0,
        "precioUnitario": precio_unitario if precio_unitario else importe_total,
        "importeItem": importe_total,
        "importeBonificacion": importe_bonificacion,
        "importeIVA": importe_iva,
        "codigoCondicionIVA": codigo_condicion_iva,
        "unidadMedida": unidad_medida,
        "bonificacionPorcentaje": bonificacion_porcentaje,
    }

def _write_item_en_fila(ws, fila, item, tipo_factura):
    codigo = item.get("codigo")
    descripcion = item.get("descripcion")
    cantidad = item.get("cantidad", 0)
    precio_unitario = item.get("precioUnitario", 0)
    importe_item = item.get("importeItem", 0)
    importe_bonificacion = item.get("importeBonificacion", 0)
    importe_iva = item.get("importeIVA", 0)
    codigo_condicion_iva = item.get("codigoCondicionIVA")
    bonificacion_porcentaje = item.get("bonificacionPorcentaje", 0)
    unidad_medida = item.get("unidadMedida", "")

    ws[f"A{fila}"] = codigo
    ws[f"B{fila}"] = descripcion

    if _safe_text(tipo_factura).upper() in ("A", "B"):
        _set_num(ws, f"D{fila}", cantidad)
        _set_num(ws, f"G{fila}", precio_unitario)
        _set_num(ws, f"I{fila}", importe_item)
        _set_num(ws, f"H{fila}", importe_bonificacion)
        _set_num(ws, f"E{fila}", importe_iva)
        iva_txt = _codigo_iva_a_texto(codigo_condicion_iva)
        if iva_txt:
            ws[f"F{fila}"] = iva_txt
    else:
        _set_num(ws, f"D{fila}", cantidad)
        ws[f"E{fila}"] = unidad_medida
        _set_num(ws, f"F{fila}", precio_unitario)
        _set_num(ws, f"I{fila}", importe_item)
        _set_num(ws, f"G{fila}", bonificacion_porcentaje)
        _set_num(ws, f"H{fila}", importe_bonificacion)


def _fecha_full_date(v):

    coerced = _coerce_date_for_excel(v)
    if coerced is not None:
        return coerced.date().strftime("%Y-%m-%d")

    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    # YYYYMMDD
    if re.fullmatch(r"\d{8}", s):
        return f"{s[0:4]}-{s[4:6]}-{s[6:8]}"
    # YYYY-MM-DD (o con /)
    m = re.fullmatch(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", s)
    if m:
        y, mo, d = m.group(1), int(m.group(2)), int(m.group(3))
        return f"{y}-{mo:02d}-{d:02d}"
    # DD/MM/YYYY
    m = re.fullmatch(r"(\d{1,2})[-/](\d{1,2})[-/](\d{4})", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), m.group(3)
        return f"{y}-{mo:02d}-{d:02d}"
    # fallback
    try:
        # último intento: parse ISO-like
        dt = _dt.datetime.fromisoformat(s)
        return dt.date().strftime("%Y-%m-%d")
    except Exception:
        return None

def _map_tipo_cmp(tipo_nota, tipo_factura):
    """
    Mapea a códigos AFIP/ARCA:
    - Factura: A=1, B=6, C=11
    - Nota Débito: A=2, B=7, C=12
    - Nota Crédito: A=3, B=8, C=13
    """
    tipo_factura = str(tipo_factura).upper().strip()
    tipo_nota = str(tipo_nota).strip().lower()

    fact = {"A": 1, "B": 6, "C": 11}
    deb = {"A": 2, "B": 7, "C": 12}
    cred = {"A": 3, "B": 8, "C": 13}

    if tipo_nota == "factura":
        return fact.get(tipo_factura)
    if tipo_nota == "debito":
        return deb.get(tipo_factura)
    if tipo_nota == "credito":
        return cred.get(tipo_factura)
    return None

def _guess_tipo_doc_rec_y_nro(doc_nro):
    """
    Heurística simple:
    - 11 dígitos => CUIT (80)
    - 8 dígitos  => DNI (96)
    - otro       => 99
    """
    s = _solo_digitos(doc_nro)
    if not s:
        return (None, None)
    if len(s) == 11:
        return (80, int(s))
    if len(s) == 8:
        return (96, int(s))
    return (99, int(s))

def _build_qr_url(*, fecha, cuit, pto_vta, tipo_cmp, nro_cmp, importe, moneda="PES", cotizacion=1.0,
                  tipo_doc_rec=None, nro_doc_rec=None, tipo_cod_aut="E", cod_aut=None):
    payload = {
        "ver": 1,
        "fecha": fecha,
        "cuit": int(cuit),
        "ptoVta": int(pto_vta),
        "tipoCmp": int(tipo_cmp),
        "nroCmp": int(nro_cmp),
        "importe": round(float(importe), 2),
        "moneda": str(moneda),
        "ctz": round(float(cotizacion), 6),
        "tipoCodAut": str(tipo_cod_aut),
        "codAut": int(cod_aut),
    }
    if tipo_doc_rec is not None and nro_doc_rec is not None:
        payload["tipoDocRec"] = int(tipo_doc_rec)
        payload["nroDocRec"] = int(nro_doc_rec)

    b64 = base64.b64encode(json.dumps(payload, separators=(",", ":"), ensure_ascii=False).encode("utf-8")).decode("ascii")
    return f"{_QR_BASE_URL}?p={b64}"

def _qr_png_bytes(qr_url):
    if qrcode is None:
        return None
    qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=6, border=1)
    qr.add_data(qr_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    # Asegurar PIL Image real
    try:
        img = img.get_image()
    except Exception:
        pass
    buf = BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def restaurar_imagenes_desde_plantilla_excel(plantilla_path, salida_xlsx_path, hojas=("Hoja1","Hoja2","Hoja3"), skip_cells_por_hoja=None):

    skip_cells_por_hoja = skip_cells_por_hoja or {}
    excel = None
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        wb_src = excel.Workbooks.Open(os.path.abspath(plantilla_path))
        wb_dst = excel.Workbooks.Open(os.path.abspath(salida_xlsx_path))

        for sh_name in hojas:
            try:
                sh_src = wb_src.Worksheets(sh_name)
                sh_dst = wb_dst.Worksheets(sh_name)
            except Exception:
                continue

            skip = set(skip_cells_por_hoja.get(sh_name, set()))

            # Copiamos shapes (imágenes, etc.)
            for shp in list(sh_src.Shapes):
                try:
                    tl = shp.TopLeftCell.Address(False, False)  # "A39"
                except Exception:
                    tl = None
                if tl and tl in skip:
                    continue
                try:
                    shp.Copy()
                    sh_dst.Paste()
                    new_shp = sh_dst.Shapes(sh_dst.Shapes.Count)
                    new_shp.Left = shp.Left
                    new_shp.Top = shp.Top
                    new_shp.Width = shp.Width
                    new_shp.Height = shp.Height
                    try:
                        new_shp.PrintObject = True
                    except Exception:
                        pass
                except Exception:
                    continue

        wb_dst.Save()
        wb_dst.Close(SaveChanges=True)
        wb_src.Close(SaveChanges=False)
    finally:
        if excel is not None:
            excel.Quit()

def insertar_qr_en_excel(salida_xlsx_path, qr_png_bytes, hojas=("Hoja1","Hoja2","Hoja3"), anchor_cell="A39",
                         width=110, height=110):

    if not qr_png_bytes:
        return

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
    try:
        tmp.write(qr_png_bytes)
        tmp.close()

        excel = None
        try:
            excel = win32.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False

            wb = excel.Workbooks.Open(os.path.abspath(salida_xlsx_path))

            for sh_name in hojas:
                try:
                    ws = wb.Worksheets(sh_name)
                except Exception:
                    continue

                # Borra cualquier shape anclada en anchor_cell (placeholder QR)
                try:
                    for shp in list(ws.Shapes):
                        try:
                            tl = shp.TopLeftCell.Address(False, False)
                        except Exception:
                            tl = None
                        if tl == anchor_cell:
                            try:
                                shp.Delete()
                            except Exception:
                                pass
                except Exception:
                    pass

                # Inserta QR
                try:
                    rng = ws.Range(anchor_cell)
                    left, top = rng.Left, rng.Top
                    pic = ws.Shapes.AddPicture(tmp.name, False, True, left, top, width, height)
                    try:
                        pic.PrintObject = True
                    except Exception:
                        pass
                except Exception:
                    continue

            wb.Save()
            wb.Close(SaveChanges=True)
        finally:
            if excel is not None:
                excel.Quit()
    finally:
        try:
            os.remove(tmp.name)
        except Exception:
            pass

def post_procesar_imagenes_y_qr(plantilla_path, factura_output, *, config, tipo_nota, tipo_factura, comprobante, validacion):

    try:
        restaurar_imagenes_desde_plantilla_excel(
            plantilla_path,
            factura_output,
            skip_cells_por_hoja={"Hoja1": {"A39"}, "Hoja2": {"A39"}, "Hoja3": {"A39"}}
        )
        escribir_log(f"{obtener_timestamp()} - Imágenes del template restauradas correctamente.")
    except Exception as e:
        escribir_log(f"{obtener_timestamp()} - WARNING: No se pudieron restaurar imágenes del template: {e}")

    try:
        fields = _resolve_comprobante_fields(comprobante, tipo_nota, tipo_factura)

        cuit = _to_int_safe(config.get("Cuit") or config.get("CUIT") or config.get("cuit"))
        pto_vta = _to_int_safe(config.get("Punto Venta") or config.get("PtoVta") or config.get("ptoVta"))
        tipo_cmp = _map_tipo_cmp(tipo_nota, tipo_factura)

        fecha_emision = _fecha_full_date(fields.get("fecha_emision"))
        doc_nro = fields.get("documento")
        importe = _to_float_safe(fields.get("importe_total"), default=0.0)
        nro_cmp = _to_int_safe(validacion[3])
        cod_aut = _to_int_safe(validacion[1])

        if not all([cuit, pto_vta, tipo_cmp, nro_cmp, cod_aut, fecha_emision]):
            escribir_log(
                f"{obtener_timestamp()} - WARNING: No se pudo armar QR (faltan datos). "
                f"cuit={cuit} ptoVta={pto_vta} tipoCmp={tipo_cmp} nroCmp={nro_cmp} codAut={cod_aut} fecha={fecha_emision}"
            )
            return

        tipo_doc_rec, nro_doc_rec = _guess_tipo_doc_rec_y_nro(doc_nro)

        qr_url = _build_qr_url(
            fecha=fecha_emision,
            cuit=cuit,
            pto_vta=pto_vta,
            tipo_cmp=tipo_cmp,
            nro_cmp=nro_cmp,
            importe=importe,
            moneda="PES",
            cotizacion=1.0,
            tipo_doc_rec=tipo_doc_rec,
            nro_doc_rec=nro_doc_rec,
            tipo_cod_aut="E",
            cod_aut=cod_aut
        )

        png = _qr_png_bytes(qr_url)
        if not png:
            escribir_log(f"{obtener_timestamp()} - WARNING: No se pudo generar PNG de QR (qrcode no instalado o falló).")
            return

        insertar_qr_en_excel(factura_output, png)
        escribir_log(f"{obtener_timestamp()} - QR ARCA insertado correctamente.")
    except Exception as e:
        escribir_log(f"{obtener_timestamp()} - WARNING: No se pudo insertar QR: {e}")


#! ------------------------------------------------------------------------------------------------------------------------------------------------------
#!
#! ARMAR EXCELS Y LUEGO GUARDAR COMO PDF'S
#!
#! ------------------------------------------------------------------------------------------------------------------------------------------------------

def qr_presente_en_xlsx(xlsx_path: str) -> bool:

    try:
        wb = load_workbook(xlsx_path)
        for sh in ("Hoja1", "Hoja2", "Hoja3"):
            if sh in wb.sheetnames:
                ws = wb[sh]
                imgs = getattr(ws, "_images", [])
                if imgs and len(imgs) > 0:
                    return True
        return False
    except Exception:

        return False

def completar_plantilla(input_path, plantilla_path, datos, cuerpo_solicitud, config, ListaValidacionCAE, tipo_factura, tipo_nota):
    contador_input = 2

    if isinstance(datos, list):
        datos = pd.DataFrame(datos)

    if isinstance(cuerpo_solicitud, list):
        cuerpo_solicitud = pd.DataFrame(cuerpo_solicitud)

    hoja_input_nombre = _sheet_name_input(tipo_nota, tipo_factura)

    for i, validacion in enumerate(ListaValidacionCAE):
        if validacion[0] is False:
            escribir_log(f"{obtener_timestamp()} - Salta CAE inválido en índice: {i}")
            escribir_log("--------------------------------------------------")
            escribir_log(f"{obtener_timestamp()} - Se abre el excel de facturación con el input, para cargar los datos antes de armar la plantilla")

            try:
                wb_input = load_workbook(input_path)
                escribir_log(f"{obtener_timestamp()} - Logró abrir correctamente el workbook, con la plantilla {input_path}")
            except FileNotFoundError:
                escribir_log(f"{obtener_timestamp()} - Error: El archivo de plantilla {input_path} no fue encontrado.")
                return
            except PermissionError:
                escribir_log(f"{obtener_timestamp()} - Error: No se tienen permisos para acceder al archivo de plantilla {input_path}.")
                return
            except Exception as e:
                escribir_log(f"{obtener_timestamp()} - Error al cargar la plantilla: {e}")
                return

            try:
                ws_input = wb_input[hoja_input_nombre]
            except KeyError:
                escribir_log(f"{obtener_timestamp()} - Error: La hoja {hoja_input_nombre} no existe en el input.")
                return

            ws_input[f"AD{contador_input}"] = "No realizo la Factura"
            ws_input[f"AE{contador_input}"] = "-"
            ws_input[f"AF{contador_input}"] = "-"
            ws_input[f"AG{contador_input}"] = "-"
            contador_input += 1

            wb_input.save(input_path)
            continue

        try:
            escribir_log("--------------------------------------------------")
            escribir_log(f"{obtener_timestamp()} - Se abre el excel de facturación con el input, para cargar los datos antes de armar la plantilla")

            try:
                wb_input = load_workbook(input_path)
                escribir_log(f"{obtener_timestamp()} - Logró abrir correctamente el workbook, con la plantilla {input_path}")
            except FileNotFoundError:
                escribir_log(f"{obtener_timestamp()} - Error: El archivo de plantilla {input_path} no fue encontrado.")
                return
            except PermissionError:
                escribir_log(f"{obtener_timestamp()} - Error: No se tienen permisos para acceder al archivo de plantilla {input_path}.")
                return
            except Exception as e:
                escribir_log(f"{obtener_timestamp()} - Error al cargar la plantilla: {e}")
                return

            try:
                ws_input = wb_input[hoja_input_nombre]
            except KeyError:
                escribir_log(f"{obtener_timestamp()} - Error: La hoja {hoja_input_nombre} no existe en el input.")
                return

            ws_input[f"AD{contador_input}"] = "Realizo la Factura"
            ws_input[f"AE{contador_input}"] = validacion[1]
            ws_input[f"AF{contador_input}"] = validacion[2]
            ws_input[f"AG{contador_input}"] = validacion[3]
            contador_input += 1

            wb_input.save(input_path)

            escribir_log("--------------------------------------------------")
            escribir_log("Iniciando proceso para completar la plantilla...")
            escribir_log(f"{obtener_timestamp()} - Ruta de la plantilla: {plantilla_path}")
            escribir_log(f"{obtener_timestamp()} - Tipo de comprobante: {tipo_nota} {tipo_factura}")

            solicitud = cuerpo_solicitud.iloc[i]
            comprobante = datos.iloc[i]
            fields = _resolve_comprobante_fields(comprobante, tipo_nota, tipo_factura)

            escribir_log(f"{obtener_timestamp()} - Logró completar los datos de solicitud y comprobante")

            if not plantilla_path:
                escribir_log(f"{obtener_timestamp()} - Error: La ruta de la plantilla no está definida.")
                return

            try:
                wb = load_workbook(plantilla_path)
                escribir_log(f"{obtener_timestamp()} - Logró abrir correctamente el workbook, con la plantilla {plantilla_path}")
            except FileNotFoundError:
                escribir_log(f"{obtener_timestamp()} - Error: El archivo de plantilla {plantilla_path} no fue encontrado.")
                return
            except PermissionError:
                escribir_log(f"{obtener_timestamp()} - Error: No se tienen permisos para acceder al archivo de plantilla {plantilla_path}.")
                return
            except Exception as e:
                escribir_log(f"{obtener_timestamp()} - Error al cargar la plantilla: {e}")
                return

            hojas = ["Hoja1", "Hoja2", "Hoja3"]

            for hoja_nombre in hojas:
                try:
                    ws = wb[hoja_nombre]

                    pto_vta_raw = config.get("Punto Venta", "-")
                    pto_vta_int = _to_int_safe(pto_vta_raw)
                    pto_vta_str = f"{pto_vta_int:05d}" if pto_vta_int is not None else str(pto_vta_raw)

                    nro_cmp_int = _to_int_safe(validacion[3])
                    nro_cmp_str = f"{nro_cmp_int:08d}" if nro_cmp_int is not None else str(validacion[3])

                    razon_social = str(config.get("Razón Social", "-") or "-")
                    domicilio_comercial = str(config.get("Domicilio Comercial", "-") or "-")
                    cuit_emisor = str(config.get("Cuit", "-") or "-")
                    ingresos_brutos = str(config.get("Ingresos Brutos") or config.get("ingresos_brutos") or "-")
                    fecha_inicio_actividades = str(
                        config.get("Fecha de Inicio de Actividades")
                        or config.get("fecha_inicio_actividades")
                        or "-"
                    )

                    ws["G3"] = pto_vta_str
                    ws["G5"] = cuit_emisor
                    ws["B4"] = razon_social
                    ws["C5"] = domicilio_comercial
                    ws["C6"] = ingresos_brutos
                    ws["C7"] = config.get("Condición IVA", "-")
                    ws["I6"] = fecha_inicio_actividades
                    ws["I3"] = nro_cmp_str

                    try:
                        rs_up = razon_social.upper()
                        a2_val = ws["A2"].value
                        if a2_val is None or str(a2_val).strip() == "":
                            ws["A2"] = rs_up
                        else:
                            if rs_up not in str(a2_val).upper():
                                ws["A2"] = f"{str(a2_val)} - {rs_up}"
                        ws["A2"].font = ws["A2"].font.copy(sz=10)
                    except Exception:
                        pass

                    try:
                        ws["C5"].font = ws["C5"].font.copy(sz=7)
                    except Exception:
                        pass

                    ws["H38"] = validacion[1]
                    ws["H39"] = validacion[2]

                    fecha_emision = fields.get("fecha_emision")
                    periodo_desde = fields.get("periodo_desde")
                    periodo_hasta = fields.get("periodo_hasta")
                    fecha_vencimiento_pago = fields.get("fecha_vencimiento_pago") or fecha_emision

                    _set_excel_date(ws, "G4", fecha_emision)
                    _set_excel_date(ws, "C8", periodo_desde)
                    _set_excel_date(ws, "E8", periodo_hasta)
                    _set_excel_date(ws, "I8", fecha_vencimiento_pago)

                    ws["C11"] = _safe_text(fields.get("condicion_iva"))
                    ws["F9"] = _safe_text(fields.get("cliente"))
                    ws["B9"] = _safe_text(fields.get("documento"))
                    ws["F11"] = _safe_text(fields.get("domicilio"))
                    ws["C13"] = _safe_text(fields.get("condicion_venta"))

                    fila = 15
                    items = _extract_array_items(solicitud)
                    if not items:
                        items = [_build_single_item_from_comprobante(comprobante, tipo_nota, tipo_factura)]

                    for item in items:
                        _write_item_en_fila(ws, fila, item, tipo_factura)
                        fila += 1

                except KeyError:
                    escribir_log(f"{obtener_timestamp()} - Error: La hoja {hoja_nombre} no existe en el archivo de plantilla.")
                except Exception as e:
                    escribir_log(f"{obtener_timestamp()} - Error al intentar completar la hoja {hoja_nombre}: {e}")

            try:
                cuit_arch = _solo_digitos(config.get("Cuit", ""))
                tipo_cmp = _map_tipo_cmp(tipo_nota, tipo_factura)
                tipo_cmp_str = f"{int(tipo_cmp):03d}" if tipo_cmp is not None else "000"

                pto_vta_int = _to_int_safe(config.get("Punto Venta"))
                pto_vta_str = f"{pto_vta_int:05d}" if pto_vta_int is not None else str(config.get("Punto Venta", "-"))

                nro_cmp_int = _to_int_safe(validacion[3])
                nro_cmp_str = f"{nro_cmp_int:08d}" if nro_cmp_int is not None else str(validacion[3])

                nombre_base = f"{cuit_arch}_{tipo_cmp_str}_{pto_vta_str}_{nro_cmp_str}"
            except Exception:
                nombre_base = f"Comprobante_{validacion[3]}"

            factura_output = os.path.join(os.path.dirname(plantilla_path), f"{nombre_base}.xlsx")
            wb.save(factura_output)
            escribir_log(f"{obtener_timestamp()} - Factura generada y guardada en {factura_output}.")

            try:
                post_procesar_imagenes_y_qr(
                    plantilla_path,
                    factura_output,
                    config=config,
                    tipo_nota=tipo_nota,
                    tipo_factura=tipo_factura,
                    comprobante=comprobante,
                    validacion=validacion
                )
            except Exception as e:
                escribir_log(f"{obtener_timestamp()} - WARNING: Falló el post-procesado de imágenes/QR: {e}")

            try:
                if not qr_presente_en_xlsx(factura_output):
                    asegurar_qr_en_factura(
                        factura_output,
                        config=config,
                        solicitud=solicitud,
                        validacion=validacion,
                    )
            except Exception as e:
                escribir_log(f"{obtener_timestamp()} - WARNING: Falló el reintento del QR: {e}")

        except Exception as e:
            escribir_log(f"{obtener_timestamp()} - Error al completar plantilla para índice {i}: {str(e)}")
            print(f"{obtener_timestamp()} - Error al completar plantilla para índice {i}: {str(e)}")

    os.remove(plantilla_path)

#! FUNCIONALIDAD PARA ARMAR PDF's Y LUEGO ALMACENARLOS EN LA CARPETA CORRESPONDIENTE
def convertir_xlsx_a_pdf(carpeta_origen, carpeta_destino):

    carpeta_dia = os.path.join(carpeta_destino, _dt.datetime.now().strftime("%d-%m-%Y"))
    os.makedirs(carpeta_dia, exist_ok=True)

    excel = None

    print(f"{obtener_timestamp()} - [PDF] Iniciando convertir_xlsx_a_pdf()")
    print(f"{obtener_timestamp()} - [PDF] carpeta_origen: {carpeta_origen}")
    print(f"{obtener_timestamp()} - [PDF] carpeta_destino: {carpeta_destino}")
    print(f"{obtener_timestamp()} - [PDF] carpeta_dia: {carpeta_dia}")
    escribir_log(f"{obtener_timestamp()} - [PDF] Iniciando convertir_xlsx_a_pdf()")
    escribir_log(f"{obtener_timestamp()} - [PDF] carpeta_origen: {carpeta_origen}")
    escribir_log(f"{obtener_timestamp()} - [PDF] carpeta_dia: {carpeta_dia}")

    pythoncom.CoInitialize()

    try:
        print(f"{obtener_timestamp()} - [PDF] Creando instancia Excel (DispatchEx)...")
        escribir_log(f"{obtener_timestamp()} - [PDF] Creando instancia Excel (DispatchEx)...")

        excel = win32.DispatchEx('Excel.Application')

        print(f"{obtener_timestamp()} - [PDF] Excel creado OK")
        escribir_log(f"{obtener_timestamp()} - [PDF] Excel creado OK")

        try:
            excel.Visible = False
            print(f"{obtener_timestamp()} - [PDF] excel.Visible=False OK")
            escribir_log(f"{obtener_timestamp()} - [PDF] excel.Visible=False OK")
        except Exception as e:
            print(f"{obtener_timestamp()} - [PDF] WARNING: No se pudo setear excel.Visible (sigo igual): {e}")
            escribir_log(f"{obtener_timestamp()} - [PDF] WARNING: No se pudo setear excel.Visible (sigo igual): {e}")

        try:
            excel.DisplayAlerts = False
            print(f"{obtener_timestamp()} - [PDF] excel.DisplayAlerts=False OK")
        except Exception as e:
            print(f"{obtener_timestamp()} - [PDF] WARNING: No se pudo setear DisplayAlerts: {e}")

        try:
            excel.ScreenUpdating = False
        except Exception:
            pass
        try:
            excel.EnableEvents = False
        except Exception:
            pass

        archivos = []
        try:
            archivos = os.listdir(carpeta_origen)
        except Exception as e:
            print(f"{obtener_timestamp()} - [PDF] ERROR: No se pudo listar carpeta_origen: {e}")
            escribir_log(f"{obtener_timestamp()} - [PDF] ERROR: No se pudo listar carpeta_origen: {e}")
            return

        print(f"{obtener_timestamp()} - [PDF] Archivos encontrados: {len(archivos)}")
        escribir_log(f"{obtener_timestamp()} - [PDF] Archivos encontrados: {len(archivos)}")

        for archivo in archivos:
            if not archivo.lower().endswith('.xlsx'):
                continue
            if archivo.startswith('~$'):
                continue

            ruta_xlsx = os.path.join(carpeta_origen, archivo)
            ruta_pdf = os.path.join(carpeta_dia, os.path.splitext(archivo)[0] + '.pdf')

            print(f"{obtener_timestamp()} - [PDF] ----")
            print(f"{obtener_timestamp()} - [PDF] Procesando XLSX: {ruta_xlsx}")
            print(f"{obtener_timestamp()} - [PDF] Destino PDF:     {ruta_pdf}")
            escribir_log(f"{obtener_timestamp()} - [PDF] Procesando XLSX: {ruta_xlsx}")
            escribir_log(f"{obtener_timestamp()} - [PDF] Destino PDF: {ruta_pdf}")

            workbook = None
            try:
                print(f"{obtener_timestamp()} - [PDF] Abriendo workbook...")
                escribir_log(f"{obtener_timestamp()} - [PDF] Abriendo workbook...")

                workbook = excel.Workbooks.Open(ruta_xlsx)

                print(f"{obtener_timestamp()} - [PDF] Workbook abierto OK, exportando a PDF...")
                escribir_log(f"{obtener_timestamp()} - [PDF] Workbook abierto OK, exportando a PDF...")

                workbook.ExportAsFixedFormat(0, ruta_pdf)

                print(f"{obtener_timestamp()} - [PDF] ExportAsFixedFormat OK, cerrando workbook...")
                escribir_log(f"{obtener_timestamp()} - [PDF] ExportAsFixedFormat OK, cerrando workbook...")

                workbook.Close(SaveChanges=False)
                workbook = None

                print(f"{obtener_timestamp()} - Convertido: {archivo} a {ruta_pdf}")
                escribir_log("")
                escribir_log(f"{obtener_timestamp()} - Convertido: {archivo} a {ruta_pdf}")

            except Exception as e:
                print(f"{obtener_timestamp()} - Error al convertir {archivo}: {e}")
                escribir_log(f"{obtener_timestamp()} - Error al convertir {archivo}: {e}")

                try:
                    if workbook is not None:
                        print(f"{obtener_timestamp()} - [PDF] Intentando cerrar workbook tras error...")
                        workbook.Close(SaveChanges=False)
                except Exception as e2:
                    print(f"{obtener_timestamp()} - [PDF] WARNING: No se pudo cerrar workbook tras error: {e2}")
                    escribir_log(f"{obtener_timestamp()} - [PDF] WARNING: No se pudo cerrar workbook tras error: {e2}")

            try:
                destino_xlsx = os.path.join(carpeta_dia, archivo)
                if os.path.exists(destino_xlsx):
                    base, ext = os.path.splitext(archivo)
                    n = 1
                    while os.path.exists(os.path.join(carpeta_dia, f"{base}_{n}{ext}")):
                        n += 1
                    destino_xlsx = os.path.join(carpeta_dia, f"{base}_{n}{ext}")

                shutil.move(ruta_xlsx, destino_xlsx)
                print(f"{obtener_timestamp()} - [PDF] Movido XLSX a: {destino_xlsx}")
                escribir_log(f"{obtener_timestamp()} - [PDF] Movido XLSX a: {destino_xlsx}")

            except Exception as e:
                print(f"{obtener_timestamp()} - [PDF] WARNING: No se pudo mover {archivo} a carpeta del día: {e}")
                escribir_log(f"{obtener_timestamp()} - WARNING: No se pudo mover {archivo} a carpeta del día: {e}")

    finally:
        try:
            if excel is not None:
                print(f"{obtener_timestamp()} - [PDF] Cerrando Excel (Quit)...")
                escribir_log(f"{obtener_timestamp()} - [PDF] Cerrando Excel (Quit)...")
                excel.Quit()
                print(f"{obtener_timestamp()} - [PDF] Excel cerrado OK")
                escribir_log(f"{obtener_timestamp()} - [PDF] Excel cerrado OK")
        except Exception as e:
            print(f"{obtener_timestamp()} - [PDF] WARNING: excel.Quit() falló: {e}")
            escribir_log(f"{obtener_timestamp()} - [PDF] WARNING: excel.Quit() falló: {e}")

        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass
