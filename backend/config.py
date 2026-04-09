import datetime
import re
import unicodedata
import pandas as pd
import openpyxl
from backend.log import obtener_timestamp


def normalizar_nombre_columna_config(value):
    if value is None:
        return ""

    text = str(value).strip()
    if not text:
        return ""

    folded = unicodedata.normalize("NFKD", text)
    folded = "".join(ch for ch in folded if not unicodedata.combining(ch))
    folded = folded.lower().replace("_", " ")
    folded = re.sub(r"[^a-z0-9]+", " ", folded)
    folded = re.sub(r"\s+", " ", folded).strip()

    aliases = {
        "cuit": "Cuit",
        "punto venta": "Punto Venta",
        "punto de venta": "Punto Venta",
        "pto vta": "Punto Venta",
        "pto venta": "Punto Venta",
        "tipo comprobante": "Tipo Comprobante",
        "numero actividad": "Numero Actividad",
        "nro actividad": "Numero Actividad",
        "con detalle de items": "¿Con detalle de Items?",
        "con detalle de item": "¿Con detalle de Items?",
        "razon social": "Razón Social",
        "domicilio comercial": "Domicilio Comercial",
        "condicion iva": "Condición IVA",
        "ingresos brutos": "Ingresos Brutos",
        "fecha de inicio de actividades": "Fecha de Inicio de Actividades",
        "fecha inicio actividades": "Fecha de Inicio de Actividades",
    }
    return aliases.get(folded, text)


def normalizar_columnas_config_df(df):
    if df is None:
        return df

    renamed = {}
    for col in df.columns:
        renamed[col] = normalizar_nombre_columna_config(col)

    df = df.rename(columns=renamed)
    if getattr(df.columns, "duplicated", None) is not None:
        df = df.loc[:, ~df.columns.duplicated(keep="first")]
    return df


def normalizar_config_dict(config):
    if not isinstance(config, dict):
        return config

    normalized = {}
    for key, value in config.items():
        canonical_key = normalizar_nombre_columna_config(key)
        existing = normalized.get(canonical_key)
        if existing in (None, ""):
            normalized[canonical_key] = value

    return normalized


def completar_campos_fijos_desde_hoja_config(df, ruta_excel, sheet_name):
    if df is None:
        return df

    try:
        wb = openpyxl.load_workbook(ruta_excel, data_only=True)
        if sheet_name not in wb.sheetnames:
            return df

        ws = wb[sheet_name]
        ingresos_brutos_fijo = ws["I2"].value
        fecha_inicio_fija = ws["J2"].value
        
        debug_msg = (
            f"{obtener_timestamp()} - [COMPLETAR_CAMPOS] Sheet={sheet_name} | "
            f"I2 (Ingresos Brutos)={repr(ingresos_brutos_fijo)} | "
            f"J2 (Fecha Inicio)={repr(fecha_inicio_fija)}"
        )
        print(debug_msg)
        from backend.log import escribir_log as elog
        try:
            elog(debug_msg)
        except:
            pass

        if "Ingresos Brutos" not in df.columns:
            df["Ingresos Brutos"] = ""
        if "Fecha de Inicio de Actividades" not in df.columns:
            df["Fecha de Inicio de Actividades"] = ""

        if ingresos_brutos_fijo not in (None, ""):
            mask_ib = df["Ingresos Brutos"].isna() | (df["Ingresos Brutos"].astype(str).str.strip() == "")
            df.loc[mask_ib, "Ingresos Brutos"] = ingresos_brutos_fijo

        if fecha_inicio_fija not in (None, ""):
            mask_fi = df["Fecha de Inicio de Actividades"].isna() | (df["Fecha de Inicio de Actividades"].astype(str).str.strip() == "")
            df.loc[mask_fi, "Fecha de Inicio de Actividades"] = fecha_inicio_fija

        return df
    except Exception:
        return df

"""
#? Detecta el formato de la fecha de manera flexible y devuelve un objeto datetime.
#? Intenta con múltiples formatos de fecha comunes.
"""
def detectar_formato_fecha(fecha):
    formatos_fecha = ['%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d', '%d-%m-%Y']
    for formato in formatos_fecha:
        try:
            return datetime.datetime.strptime(fecha, formato)
        except ValueError:
            continue
    raise ValueError(f"{obtener_timestamp()} - Formato de fecha no reconocido.")

#? Función para validar y transformar la fecha según el servicio
def validar_y_transformar_fecha(fecha, servicio):
    try:
        # Detectar y convertir la fecha si es cadena
        if isinstance(fecha, str):
            fecha_dt = detectar_formato_fecha(fecha)
        else:
            # Si ya es un objeto datetime
            fecha_dt = pd.to_datetime(fecha, errors='coerce')
        
        if pd.isnull(fecha_dt):
            raise ValueError("Fecha inválida")

        # Devolver en el formato correspondiente según el servicio
        if servicio == 'MTXCA':
            return fecha_dt.strftime('%Y-%m-%d')  # Formato YYYY-MM-DD
        elif servicio == 'FEV1':
            return fecha_dt.strftime('%Y%m%d')    # Formato YYYYMMDD
        else:
            raise ValueError("Servicio inválido. Utiliza 'MTXCA' o 'FEV1'.")
    
    except Exception as e:
        print(f"{obtener_timestamp()} - Error al procesar la fecha: {e}")
        return fecha

#? Eliminar ".0" de números enteros
def remove_dot_zero(value):
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _normalizar_valor_config(value):
    if value is None:
        return ""
    if pd.isna(value):
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (datetime.datetime, datetime.date, pd.Timestamp)):
        return value.strftime('%d/%m/%Y')
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _norm_int(v):
    """Normaliza un valor a entero, retorna None si está vacío/None."""
    try:
        if v is None:
            return None
        if isinstance(v, str):
            v = v.strip()
            if v == "":
                return None
        return int(float(v))
    except Exception:
        return None


def _norm_text(v):
    """Normaliza un valor a texto limpio."""
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _si_no_to_bool(v):
    """Convierte si/no a booleano."""
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    return str(v).strip().lower() in ("si", "sí", "true", "1", "y", "yes")


def _build_header_map(ws):
    """Crea un mapeo (nombre_normalizado -> índice_columna) desde la primera fila."""
    hmap = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value is None:
            continue
        key = normalizar_nombre_columna_config(cell.value)
        if key:
            hmap[key] = col_idx
    return hmap


def _find_row_by_cuit(ws, col_cuit, cuit):
    """Busca la fila que contiene el CUIT especificado (1-based row index o None)."""
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=col_cuit).value
        if _norm_int(val) == cuit:
            return r
    return None


def leer_config_por_cuit_desde_hoja(ruta_excel, sheet_name, cuit):
    """
    Lee una configuración de Config.xlsx de una hoja específica por CUIT.
    Normaliza todos los valores según tipo (int, bool, texto, fecha).
    Aplicar fallback I2/J2 para campos especiales.
    
    Retorna diccionario con clave normalizada (ej: "Razón Social", "Ingresos Brutos")
    o None si no encuentra el CUIT.
    """
    try:
        if not openpyxl.utils.get_column_letter:
            pass  # Asegura que openpyxl está disponible
        
        wb = openpyxl.load_workbook(ruta_excel, data_only=True)
        if sheet_name not in wb.sheetnames:
            return None
        
        ws = wb[sheet_name]
        hmap = _build_header_map(ws)
        
        col_cuit = hmap.get("Cuit")
        if not col_cuit:
            return None
        
        row = _find_row_by_cuit(ws, col_cuit, cuit)
        if row is None:
            return None
        
        def get(header_name):
            """Obtiene valor de una columna por su nombre normalizado."""
            col_idx = hmap.get(header_name)
            return ws.cell(row=row, column=col_idx).value if col_idx else None
        
        # Leer con fallback I2/J2 para campos especiales
        ingresos_brutos = get("Ingresos Brutos")
        if ingresos_brutos in (None, ""):
            ingresos_brutos = ws["I2"].value
        
        fecha_inicio_actividades = get("Fecha de Inicio de Actividades")
        if fecha_inicio_actividades in (None, ""):
            fecha_inicio_actividades = ws["J2"].value
        
        # Retornar diccionario con valores normalizados por tipo
        return {
            "Cuit": _norm_int(get("Cuit")),
            "Punto Venta": _norm_int(get("Punto Venta")),
            "Tipo Comprobante": _norm_int(get("Tipo Comprobante")),
            "Numero Actividad": _norm_int(get("Numero Actividad")),
            "¿Con detalle de Items?": _si_no_to_bool(get("¿Con detalle de Items?")),
            "Razón Social": _norm_text(get("Razón Social")),
            "Domicilio Comercial": _norm_text(get("Domicilio Comercial")),
            "Condición IVA": _norm_text(get("Condición IVA")),
            "Ingresos Brutos": _norm_text(_normalizar_valor_config(ingresos_brutos)),
            "Fecha de Inicio de Actividades": _norm_text(_normalizar_valor_config(fecha_inicio_actividades)),
        }
    except Exception as e:
        print(f"{obtener_timestamp()} - Error en leer_config_por_cuit_desde_hoja: {e}")
        return None


def normalizar_config_para_pdf(config):
    if not isinstance(config, dict):
        return config

    config_normalizada = normalizar_config_dict(config)

    for key in (
        "Cuit",
        "Punto Venta",
        "Tipo Comprobante",
        "Numero Actividad",
        "¿Con detalle de Items?",
        "Razón Social",
        "Domicilio Comercial",
        "Condición IVA",
    ):
        if key in config_normalizada:
            config_normalizada[key] = _normalizar_valor_config(config_normalizada.get(key))

    ingresos_brutos = (
        config_normalizada.get("Ingresos Brutos")
        or config_normalizada.get("ingresos_brutos")
        or ""
    )
    fecha_inicio_actividades = (
        config_normalizada.get("Fecha de Inicio de Actividades")
        or config_normalizada.get("fecha_inicio_actividades")
        or ""
    )

    ingresos_brutos = _normalizar_valor_config(ingresos_brutos)
    fecha_inicio_actividades = _normalizar_valor_config(fecha_inicio_actividades)

    config_normalizada["Ingresos Brutos"] = ingresos_brutos
    config_normalizada["ingresos_brutos"] = ingresos_brutos
    config_normalizada["Fecha de Inicio de Actividades"] = fecha_inicio_actividades
    config_normalizada["fecha_inicio_actividades"] = fecha_inicio_actividades

    return config_normalizada

#? Función para obtener los datos de cada hoja de Excel por separado
def leer_datos_excel(ruta_archivo, tipo_factura):
    wb = openpyxl.load_workbook(ruta_archivo)
    
    datos_factura = None
    datos_nota_credito = None
    datos_nota_debito = None

    hoja_leer1 = 'Factura {}'.format(tipo_factura)
    hoja_leer2 = 'Nota Credito {}'.format(tipo_factura) 
    hoja_leer3 = 'Nota Debito {}'.format(tipo_factura)

    print(hoja_leer1)
    print(hoja_leer2)
    print(hoja_leer3)

    # Leer Hoja1 (Facturas)
    if hoja_leer1 in wb.sheetnames:
        try:
            hoja_factura = wb[hoja_leer1]
        except:
            print("No se encontró la facturación general de Factura {}, para realizar la lectura de datos".format(tipo_factura))
        # Leer todas las filas de Hoja1 (ignorando el encabezado)
        datos_factura = [
            [celda if celda is not None else "" for celda in fila]
            for fila in hoja_factura.iter_rows(min_row=2, values_only=True)  # Ignorar encabezado
        ]
        # Convertir en DataFrame solo si hay datos
        if len(datos_factura) > 0:
            columnas_factura = [celda.value for celda in hoja_factura[1]]  # Tomar los encabezados de la primera fila
            datos_factura = pd.DataFrame(datos_factura, columns=columnas_factura)
        else:
            datos_factura = None

    # Leer Hoja2 (Nota de Crédito)
    if hoja_leer2 in wb.sheetnames:
        try:
            hoja_nota_credito = wb[hoja_leer2]
        except:
            print("No se encontró la Nota de Credito {}, para realizar la lectura de datos".format(tipo_factura))
        # Leer todas las filas de Hoja2 (ignorando el encabezado)
        datos_nota_credito = [
            [celda if celda is not None else "" for celda in fila]
            for fila in hoja_nota_credito.iter_rows(min_row=2, values_only=True)  # Ignorar encabezado
        ]
        # Convertir en DataFrame solo si hay datos
        if len(datos_nota_credito) > 0:
            columnas_nota_credito = [celda.value for celda in hoja_nota_credito[1]]  # Tomar los encabezados de la primera fila
            datos_nota_credito = pd.DataFrame(datos_nota_credito, columns=columnas_nota_credito)
        else:
            datos_nota_credito = None

    # Leer Hoja3 (Nota de Debito)
    if hoja_leer3 in wb.sheetnames:
        try:
            hoja_nota_debito = wb[hoja_leer3]
        except:
            print("No se encontró la Nota de Debito {}, para realizar la lectura de datos".format(tipo_factura))
        # Leer todas las filas de Hoja3 (ignorando el encabezado)
        datos_nota_debito = [
            [celda if celda is not None else "" for celda in fila]
            for fila in hoja_nota_debito.iter_rows(min_row=2, values_only=True)  # Ignorar encabezado
        ]
        # Convertir en DataFrame solo si hay datos
        if len(datos_nota_debito) > 0:
            columnas_nota_debito = [celda.value for celda in hoja_nota_debito[1]]  # Tomar los encabezados de la primera fila
            datos_nota_debito = pd.DataFrame(datos_nota_debito, columns=columnas_nota_debito)
        else:
            datos_nota_debito = None

    print(datos_factura)

    return datos_factura, datos_nota_credito, datos_nota_debito
