import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter import ttk
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from shutil import copyfile
import os
import openpyxl
import re
import unicodedata

# * EXCEL Y DISPLAY DE LAS COLUMNAS:
from categorias.diccionarios import columns_factura_a, columns_factura_a_credito, columns_factura_a_debito
from categorias.diccionarios import columns_factura_b, columns_factura_b_credito, columns_factura_b_debito
from categorias.diccionarios import columns_factura_c, columns_factura_c_credito, columns_factura_c_debito
from categorias.diccionarios import opciones_iva, opciones_tipo_comprobante, doc_tipo, opciones_codigo_iva 
from categorias.diccionarios import opciones_codigo_tributos, calculo_codigo_iva, opciones_codigo_concepto

# Definición de las plantillas
templates = {
    "Factura A": "vacio_factura_a.xlsx",
    "Nota Credito A": "vacio_factura_a.xlsx",
    "Nota Debito A": "vacio_factura_a.xlsx",
    "Factura B": "vacio_factura_b.xlsx",
    "Nota Credito B": "vacio_factura_b.xlsx",
    "Nota Debito B": "vacio_factura_b.xlsx",
    "Factura C": "vacio_factura_c.xlsx",
    "Nota Credito C": "vacio_factura_c.xlsx",
    "Nota Debito C": "vacio_factura_c.xlsx",
}

# Datos iniciales
tab_names = [
    "Factura A", "Nota Credito A", "Nota Debito A",
    "Factura B", "Nota Credito B", "Nota Debito B",
    "Factura C", "Nota Credito C", "Nota Debito C"
]

# Variable global para almacenar el archivo seleccionado
selected_file = None  
data_entries = {tab_name: [] for tab_name in tab_names}
current_workbook = None


def normalizar_texto(texto):
    """Normaliza un texto eliminando tildes y convirtiendo a minúsculas."""
    return unicodedata.normalize("NFKD", texto).encode("ASCII", "ignore").decode("ASCII").strip().lower()

def buscar_clave_parcial(clave_buscada, diccionario):
    """Busca una clave en el diccionario que contenga `clave_buscada` (ignorando corchetes y espacios)."""
    clave_buscada = normalizar_texto(clave_buscada)
    
    for clave in diccionario.keys():
        clave_normalizada = normalizar_texto(re.sub(r"\[.*?\]", "", clave))  # Elimina texto entre corchetes
        if clave_buscada in clave_normalizada:
            return clave  # Retorna la clave encontrada en el diccionario

def select_existing_file():
    """Permite al usuario seleccionar un archivo de facturación existente."""
    global selected_file, current_workbook

    file_path = filedialog.askopenfilename(
        title="Seleccionar archivo de facturación",
        filetypes=[("Excel files", "*.xlsx")],
        initialdir="input/"
    )

    if not file_path:
        return  # Si el usuario cancela, no hace nada

    try:
        current_workbook = openpyxl.load_workbook(file_path)
        selected_file = file_path  # Guarda la ruta del archivo seleccionado
        messagebox.showinfo("Archivo seleccionado", f"Se cargará la información en:\n{file_path}")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo abrir el archivo seleccionado: {e}")


def load_template(tab_name):
    """Cargar la plantilla de Excel según el tipo de facturación."""
    global current_workbook
    if "Factura A" in tab_name or "Nota Credito A" in tab_name or "Nota Debito A" in tab_name:
        current_workbook = load_workbook("vacio_factura_a.xlsx")
    elif "Factura B" in tab_name or "Nota Credito B" in tab_name or "Nota Debito B" in tab_name:
        current_workbook = load_workbook("vacio_factura_b.xlsx")
    elif "Factura C" in tab_name or "Nota Credito C" in tab_name or "Nota Debito C" in tab_name:
        current_workbook = load_workbook("vacio_factura_c.xlsx")
    else:
        messagebox.showerror("Error", f"No se encontró una plantilla para {tab_name}")
        return


def find_first_empty_row(sheet):
    """Encuentra la primera fila completamente vacía en una hoja de Excel."""
    for row in range(2, sheet.max_row + 2):  # Empieza en la fila 2
        if all(sheet[f"{get_column_letter(col)}{row}"].value is None for col in range(1, sheet.max_column + 1)):
            return row
    return sheet.max_row + 1  # Si no encontró vacías, agrega al final


def save_to_excel():
    """Guardar los datos en el libro actual."""
    global current_workbook, selected_file
    if current_workbook is None:
        messagebox.showerror("Error", "No hay un archivo cargado. Seleccione un archivo o cree uno nuevo.")
        return

    hojas_modificadas = set()  # Para rastrear las hojas modificadas

    for tab_name, entries in data_entries.items():
        if not entries:
            continue

        try:
            # Seleccionar la hoja adecuada o crearla si no existe
            sheet = current_workbook[tab_name] if tab_name in current_workbook.sheetnames else current_workbook.create_sheet(tab_name)
            
            columns = (
                columns_factura_a["excel"] if "Factura A" in tab_name else
                columns_factura_b["excel"] if "Factura B" in tab_name else
                columns_factura_c["excel"]
            )

            if tab_name in ["Factura A", "Nota Credito A", "Nota Debito A"]:
                guardar_en = "Factura A"
            elif tab_name in ["Factura B", "Nota Credito B", "Nota Debito B"]:
                guardar_en = "Factura B"
            elif tab_name in ["Factura C", "Nota Credito C", "Nota Debito C"]:
                guardar_en = "Factura C"

            # Escribir los datos en la hoja
            for row_index, entry in enumerate(entries, start=find_first_empty_row(sheet)):
                for col_index, value in enumerate(entry, start=1):
                    sheet[f"{get_column_letter(col_index)}{row_index}"] = value
            
            hojas_modificadas.add(guardar_en)

        except Exception as e:
            messagebox.showerror("Error", f"Error al escribir en la hoja '{tab_name}': {e}")

    try:
        # Guardar el archivo seleccionado si existe, o crear uno nuevo si no
        if selected_file:
            save_path = selected_file
        else:
            save_path = f"input/{guardar_en}/Facturacion.xlsx"
        
        current_workbook.save(save_path)
        messagebox.showinfo("Éxito", f"Datos guardados en {save_path}")

    except Exception as e:
        messagebox.showerror("Error", f"Error al guardar el archivo: {e}")

def add_entry(tab_name, entry_data, entries, columns):
    """Agregar una nueva entrada a la lista de datos y limpiar las entradas."""
    if not any(entry_data):  # Verifica que no esté vacío
        messagebox.showerror("Error", "No se pueden agregar datos vacíos.")
        return
    
    if tab_name not in data_entries:
        data_entries[tab_name] = []

    # Convertir el texto de "Código Otros Tributos" a su valor numérico
    codigo_tributo_text = entries["Codigo Otros Tributos [Código del tributo aplicado]"].get()
    codigo_tributo_value = opciones_codigo_tributos.get(codigo_tributo_text, 99)  # 99 es "Otros" por defecto

    # Encontrar el índice de "Código Otros Tributos" en las columnas
    tributo_index = columns["display"].index("Codigo Otros Tributos [Código del tributo aplicado]")

    # Reemplazar el valor de la entrada en el índice correspondiente
    entry_data[tributo_index] = codigo_tributo_value

    # Convertir "Codigo Condición IVA" de texto a su valor numérico
    codigo_iva_text = entries["Codigo Condicion IVA [Identificador numérico del tipo de IVA]"].get()
    codigo_iva_value = opciones_codigo_iva.get(codigo_iva_text, 0)

    # Encontrar el índice de "Codigo Condición IVA" en las columnas
    iva_index = columns["display"].index("Codigo Condicion IVA [Identificador numérico del tipo de IVA]")
    
    # Reemplazar el valor de la entrada en el índice correspondiente
    entry_data[iva_index] = codigo_iva_value

    # Convertir "Concepto" de texto a su valor numérico
    concepto_text = entries["Concepto [Código de concepto - 1, 2 o 3]"].get()
    concepto_value = opciones_codigo_concepto.get(concepto_text, 0)

    # Encontrar el índice de "Concepto" en las columnas
    concepto_index = columns["display"].index("Concepto [Código de concepto - 1, 2 o 3]")

    # Reemplazar el valor de la entrada en el índice correspondiente
    entry_data[concepto_index] = concepto_value

    # Agregar la entrada convertida a la lista de datos
    data_entries[tab_name].append(entry_data)
    
    # Limpiar las entradas en la UI
    for entry in entries.values():
        entry.delete(0, tk.END)

    messagebox.showinfo("Éxito", f"Datos agregados a {tab_name}")


def create_tab_content(tab, tab_name, columns):
    """Genera los campos de entrada y el botón OK en una pestaña."""
    entries = {}

    def update_total(event=None):
        """Calcula el total multiplicando precio unitario por cantidad."""
        print("DEBUG - update_total() se ejecutó")  #  Debug
        try:
            # Buscar claves correctas dentro de entries
            price_key = next((k for k in entries.keys() if "Precio Unitario" in k), None)
            quantity_key = next((k for k in entries.keys() if "Cantidad" in k), None)
            total_key = next((k for k in entries.keys() if "Total" in k), None)

            if not price_key or not quantity_key or not total_key:
                print(f"DEBUG - Claves no encontradas: Precio: {price_key}, Cantidad: {quantity_key}, Total: {total_key}")
                return

            price = float(entries[price_key].get().replace(',', '.')) if entries[price_key].get() else 0
            quantity = float(entries[quantity_key].get().replace(',', '.')) if entries[quantity_key].get() else 0
            total = price * quantity
            print(f"DEBUG - Precio: {price}, Cantidad: {quantity}, Total: {total}")  #  Debug

            entries[total_key].delete(0, tk.END)
            entries[total_key].insert(0, f"{total:.2f}")
            entries[total_key].config(fg="black")

            update_iva()  # Forzar actualización de IVA
                
        except Exception as e:
            print(f"DEBUG - Error en update_total(): {e}") #  Debug


    def update_iva(event=None):
        """Calcula el importe del IVA según el total y el código de condición IVA."""
        print("DEBUG - update_iva() se ejecutó")  # Debug

        try:
            # Buscar claves correctas dentro de entries
            iva_key = next((k for k in entries.keys() if "Código Condición IVA" in k), None)
            total_key = next((k for k in entries.keys() if "Total" in k), None)
            importe_iva_key = next((k for k in entries.keys() if "Importe IVA" in k), None)

            if not iva_key or not total_key or not importe_iva_key:
                print(f"DEBUG - Claves no encontradas: IVA: {iva_key}, Total: {total_key}, Importe IVA: {importe_iva_key}")
                return

            # Obtener valores de los widgets
            iva_widget = entries[iva_key]
            total_widget = entries[total_key]
            importe_iva_widget = entries[importe_iva_key]

            if not isinstance(iva_widget, ttk.Combobox):
                print(f"DEBUG - ERROR: El widget de Código Condición IVA no es un Combobox: {type(iva_widget)}")
                return

            valor_seleccionado = iva_widget.get()
            print(f"DEBUG - Valor real obtenido del Combobox: '{valor_seleccionado}'")  # Debug

            if valor_seleccionado not in calculo_codigo_iva:
                print(f"DEBUG - Clave no encontrada en calculo_codigo_iva: '{valor_seleccionado}'")
                return

            porcentaje_iva = calculo_codigo_iva[valor_seleccionado]

            # Obtener el total
            try:
                total = float(total_widget.get().replace(',', '.')) if total_widget.get() else 0.0
            except ValueError:
                total = 0.0

            # Si total es 0, no hacer nada
            if total <= 0:
                print("DEBUG - Total es 0, no se calcula IVA")
                return

            # Calcular el importe del IVA
            importe_iva = round(total * porcentaje_iva, 2)
            print(f"DEBUG - Importe IVA calculado: {importe_iva}")  # Debug

            if isinstance(importe_iva_widget, tk.Entry):
                importe_iva_widget.config(state="normal")
                importe_iva_widget.delete(0, tk.END)
                importe_iva_widget.insert(0, f"{importe_iva:.2f}")
                importe_iva_widget.config(state="readonly")
            else:
                print(f"DEBUG - ERROR: El widget de Importe IVA no es un Entry: {type(importe_iva_widget)}")

        except Exception as e:
            print(f"DEBUG - Error en update_iva(): {e}")  # Debug


    for i, label in enumerate(columns["display"]):
        if "[" in label and "]" in label:
            parts = label.split("[", 1)
            main_text = parts[0].strip()
            extra_text = "[" + parts[1].strip()
        else:
            main_text = label
            extra_text = ""

        frame_inner = tk.Frame(tab, bg="white")
        frame_inner.grid(row=i, column=0, sticky="w")
        
        text_widget = tk.Text(frame_inner, height=1, width=65, borderwidth=0, bg="white")
        text_widget.grid(row=i, column=0, sticky="w", padx=5, pady=2)
        text_widget.insert("1.0", main_text, "main")
        if extra_text:
            text_widget.insert("end", f" {extra_text}", "extra")
        text_widget.tag_config("main", font=("Arial", 9), foreground="black")
        text_widget.tag_config("extra", font=("Arial", 8), foreground="#4d4d4d")
        text_widget.config(state="disabled")

        if label.startswith("Tipo de Dato"):
            entry = ttk.Combobox(tab, values=opciones_tipo_comprobante, width=25)
            entry.current(0)

        elif label.startswith("Concepto"):
            entry = ttk.Combobox(tab, values=list(opciones_codigo_concepto.keys()), width = 30)
            entry.current(0)

        elif label.startswith("Tipo de Documento"):
            entry = ttk.Combobox(tab, values=doc_tipo, width=30)
            entry.current(0)

        elif label.startswith("Condición frente al IVA"):
            entry = ttk.Combobox(tab, values=list(opciones_iva.keys()), width=45)
            entry.current(0)
            
        elif "Codigo Condicion IVA" in label:
            entry = ttk.Combobox(tab, values=list(opciones_codigo_iva.keys()), width=25)
            entry.current(0)
            entry.bind("<<ComboboxSelected>>", update_iva)

            if isinstance(entry, ttk.Combobox):  
                entries[label] = entry
            else:
                print(f"DEBUG - ¡Error! No se guardó correctamente el Combobox en entries[{label}]")
            
        elif label.startswith("Codigo Otros Tributos"):
            entry = ttk.Combobox(tab, values=list(opciones_codigo_tributos.keys()), width=50)
            entry.current(0)

        elif label.startswith("Fecha de emisión") or label.startswith("Periodo Desde") or label.startswith("Periodo Hasta"):
            entry = tk.Entry(tab, fg="gray", width=30)
            entry.insert(0, "Ejemplo: 2025-01-27")
            entry.bind("<FocusIn>", lambda e, entry=entry: clear_placeholder(e, entry))
            entry.bind("<FocusOut>", lambda e, entry=entry: restore_placeholder(e, entry))

        elif label.startswith("Importe") or label.startswith("Cantidad") or label.startswith("Total") or label.startswith("Precio"):
            entry = tk.Entry(tab, fg="gray", width=30)
            entry.insert(0, "0.00")
            entry.bind("<FocusIn>", lambda e, entry=entry: clear_placeholder(e, entry))
            entry.bind("<FocusOut>", lambda e, entry=entry: restore_placeholder(e, entry))
            entry.bind("<KeyRelease>", update_total)  # Agregar el evento para actualizar el total al escribir

        else:
            entry = tk.Entry(tab, fg="gray", width=30)
            entry.insert(0, "Ingresar Dato")
            entry.bind("<FocusIn>", lambda e, entry=entry: clear_placeholder(e, entry))
            entry.bind("<FocusOut>", lambda e, entry=entry: restore_placeholder(e, entry))

        entry.grid(row=i, column=1, padx=5, pady=2)
        entries[label] = entry

    # Conectar eventos SOLO si los campos existen
    if "Precio Unitario" in entries and "Cantidad" in entries and "Total" in entries:
        entries["Precio Unitario"].bind("<KeyRelease>", update_total)
        entries["Cantidad"].bind("<KeyRelease>", update_total)

    if "Importe Total" in entries and "Código Condición IVA" in entries and "Importe IVA" in entries:
        entries["Código Condición IVA"].bind("<<ComboboxSelected>>", update_iva)

    btn = tk.Button(tab, text="Agregar datos al excel",
                    font=("Arial", 11, "bold"), bg="#32CD32", fg="white",
                    relief="solid", borderwidth=2, width=22, height=2,
                    activebackground="#2BA82B", activeforeground="white",
                    command=lambda: add_entry(tab_name, [entry.get() for entry in entries.values()], entries, columns))
    btn.grid(row=i+1, column=1, padx=5, pady=5)


def clear_placeholder(event, entry):
    if entry.get() == "Ejemplo: 2025-01-27" or entry.get() == "Ingresar Dato" or entry.get() == "0.00":
        entry.delete(0, tk.END)
        entry.config(fg="black")

def restore_placeholder(event, entry):
    if not entry.get():
        entry.insert(0, "Ejemplo: 2025-01-27" if "Fecha" in entry.master.grid_info() else "Ingresar Dato")
        entry.config(fg="gray")
        
def finalize():
    """Finalizar y guardar los datos en archivos separados por tipo."""
    global current_workbook
    if not any(data_entries.values()):  # Verifica si hay datos
        messagebox.showerror("Error", "No hay datos para guardar.")
        return

    try:
        for tab_name, entries in data_entries.items():
            if not entries:
                continue

            # Cargar la plantilla correspondiente si no está cargada
            if "Factura A" in tab_name or "Nota Credito A" in tab_name or "Nota Debito A":
                load_template("Factura A")
            elif "Factura B" in tab_name or "Nota Credito B" in tab_name or "Nota Debito B":
                load_template("Factura B")
            elif "Factura C" in tab_name or "Nota Credito C" in tab_name or "Nota Debito C":
                load_template("Factura C")

            # Guardar los datos
            save_to_excel()

        messagebox.showinfo("Éxito", "Todos los archivos se han guardado correctamente.")
    except Exception as e:
        messagebox.showerror("Error", f"Error al finalizar: {e}")