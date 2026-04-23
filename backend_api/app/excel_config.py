# app/excel_config.py
import os
import re
import time
from typing import Any, Dict, List, Optional
import datetime
import unicodedata

import openpyxl

SHEET_BY_TIPO = {"A": "FacturaA", "B": "FacturaB", "C": "FacturaC"}

EXCEL_HEADERS = [
    "Cuit",
    "Punto Venta",
    "Tipo Comprobante",
    "Numero Actividad",
    "¿Con detalle de Items?",
    "Razón Social",
    "Domicilio Comercial",
    "Condición IVA",
    "Ingresos Brutos",
    "Fecha de Inicio de Actividades"
]

API_TO_EXCEL = {
    "cuit": "Cuit",
    "punto_venta": "Punto Venta",
    "tipo_comprobante": "Tipo Comprobante",
    "numero_actividad": "Numero Actividad",
    "con_detalle_items": "¿Con detalle de Items?",
    "razon_social": "Razón Social",
    "domicilio_comercial": "Domicilio Comercial",
    "condicion_iva": "Condición IVA",
    "ingresos_brutos": "Ingresos Brutos",
    "fecha_inicio_actividades": "Fecha de Inicio de Actividades"
}


def normalizar_nombre_columna_config(value: Any) -> str:
    if value is None:
        return ""

    text = str(value).strip()
    if not text:
        return ""

    folded = unicodedata.normalize("NFKD", text)
    folded = "".join(ch for ch in folded if not unicodedata.combining(ch))
    folded = folded.lower().replace("_", " ")
    folded = re.sub(r"[^a-z0-9]+", " ", folded)
    folded = re.sub(r"\s+", " ", folded).strip()

    aliases = {
        "cuit": "Cuit",
        "punto venta": "Punto Venta",
        "punto de venta": "Punto Venta",
        "pto vta": "Punto Venta",
        "pto venta": "Punto Venta",
        "tipo comprobante": "Tipo Comprobante",
        "numero actividad": "Numero Actividad",
        "nro actividad": "Numero Actividad",
        "con detalle de items": "¿Con detalle de Items?",
        "con detalle de item": "¿Con detalle de Items?",
        "razon social": "Razón Social",
        "domicilio comercial": "Domicilio Comercial",
        "condicion iva": "Condición IVA",
        "ingresos brutos": "Ingresos Brutos",
        "fecha de inicio de actividades": "Fecha de Inicio de Actividades",
        "fecha inicio actividades": "Fecha de Inicio de Actividades",
    }
    return aliases.get(folded, text)


def _normalizar_valor_config(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()

def _norm_int(v: Any) -> Optional[int]:
    try:
        if v is None:
            return None
        if isinstance(v, str):
            v = v.strip()
            if v == "":
                return None
        return int(float(v))
    except Exception:
        return None


def _bool_to_si_no(v: Any) -> str:
    if isinstance(v, str):
        vv = v.strip().lower()
        if vv in ("si", "sí", "true", "1", "y", "yes"):
            return "si"
        return "no"
    return "si" if bool(v) else "no"


def _si_no_to_bool(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    return str(v).strip().lower() in ("si", "sí", "true", "1", "y", "yes")


def _norm_text(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


class _LockFile:

    def __init__(self, lock_path: str, timeout_sec: float = 10.0):
        self.lock_path = lock_path
        self.timeout_sec = timeout_sec
        self._fd = None

    def __enter__(self):
        start = time.time()
        while True:
            try:
                self._fd = os.open(self.lock_path, os.O_CREAT | os.O_EXCL | os.O_RDWR)
                os.write(self._fd, str(os.getpid()).encode("utf-8"))
                return self
            except FileExistsError:
                if (time.time() - start) > self.timeout_sec:
                    raise TimeoutError(f"No se pudo adquirir lock: {self.lock_path}")
                time.sleep(0.05)

    def __exit__(self, exc_type, exc, tb):
        try:
            if self._fd is not None:
                os.close(self._fd)
            if os.path.exists(self.lock_path):
                os.remove(self.lock_path)
        except Exception:
            pass


def _open_wb(excel_path: str):
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"No existe Config.xlsx en: {excel_path}")
    return openpyxl.load_workbook(excel_path)


def _header_map(ws) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value is None:
            continue
        key = normalizar_nombre_columna_config(cell.value)
        if key:
            m[key] = col_idx
    return m


def _ensure_headers(ws, hmap: Dict[str, int]):
    for h in EXCEL_HEADERS:
        if h not in hmap:
            new_col = ws.max_column + 1
            ws.cell(row=1, column=new_col).value = h
            hmap[h] = new_col


def _find_row_by_cuit(ws, col_cuit: int, cuit: int) -> Optional[int]:
    """Devuelve row_idx (1-based) de la fila que matchea CUIT, o None."""
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=col_cuit).value
        if _norm_int(val) == cuit:
            return r
    return None


def _row_to_api(ws, row_idx: int, hmap: Dict[str, int]) -> Dict[str, Any]:
    def get(h: str):
        ci = hmap.get(h)
        return ws.cell(row=row_idx, column=ci).value if ci else None

    ingresos_brutos = get("Ingresos Brutos")
    if ingresos_brutos in (None, ""):
        ingresos_brutos = ws["I2"].value

    fecha_inicio_actividades = get("Fecha de Inicio de Actividades")
    if fecha_inicio_actividades in (None, ""):
        fecha_inicio_actividades = ws["J2"].value

    return {
        "cuit": _norm_int(get("Cuit")),
        "punto_venta": _norm_int(get("Punto Venta")),
        "tipo_comprobante": _norm_int(get("Tipo Comprobante")),
        "numero_actividad": _norm_int(get("Numero Actividad")),
        "con_detalle_items": _si_no_to_bool(get("¿Con detalle de Items?")),
        "razon_social": _norm_text(get("Razón Social")),
        "domicilio_comercial": _norm_text(get("Domicilio Comercial")),
        "condicion_iva": _norm_text(get("Condición IVA")),
        "ingresos_brutos": _norm_text(_normalizar_valor_config(ingresos_brutos)),
        "fecha_inicio_actividades": _norm_text(_normalizar_valor_config(fecha_inicio_actividades)),
    }

def list_cuits(excel_path: str) -> List[Dict[str, Any]]:
    lock_path = excel_path + ".lock"
    with _LockFile(lock_path):
        wb = _open_wb(excel_path)

        seen = {}
        for _, sheet in SHEET_BY_TIPO.items():
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            hmap = _header_map(ws)
            col_cuit = hmap.get("Cuit")
            if not col_cuit:
                continue

            for r in range(2, ws.max_row + 1):
                c = _norm_int(ws.cell(row=r, column=col_cuit).value)
                if not c:
                    continue
                if c not in seen:
                    rs = ""
                    col_rs = hmap.get("Razón Social")
                    if col_rs:
                        rs = ws.cell(row=r, column=col_rs).value or ""
                    seen[c] = {"cuit": c, "razon_social": rs}

        return [seen[k] for k in sorted(seen.keys())]


def get_config_by_cuit(excel_path: str, tipo: str, cuit: int) -> Dict[str, Any]:
    tipo = tipo.upper()
    if tipo not in SHEET_BY_TIPO:
        raise ValueError("tipo debe ser A, B o C")

    lock_path = excel_path + ".lock"
    with _LockFile(lock_path):
        wb = _open_wb(excel_path)
        ws = wb[SHEET_BY_TIPO[tipo]]
        hmap = _header_map(ws)
        _ensure_headers(ws, hmap)

        col_cuit = hmap["Cuit"]
        row = _find_row_by_cuit(ws, col_cuit, cuit)
        if row is None:
            raise KeyError("CUIT no encontrado")
        return _row_to_api(ws, row, hmap)


def upsert_config_by_cuit(excel_path: str, tipo: str, cuit: int, data: Dict[str, Any]) -> Dict[str, Any]:
    tipo = tipo.upper()
    if tipo not in SHEET_BY_TIPO:
        raise ValueError("tipo debe ser A, B o C")

    lock_path = excel_path + ".lock"
    with _LockFile(lock_path):
        wb = _open_wb(excel_path)
        ws = wb[SHEET_BY_TIPO[tipo]]
        hmap = _header_map(ws)
        _ensure_headers(ws, hmap)

        col_cuit = hmap["Cuit"]
        row = _find_row_by_cuit(ws, col_cuit, cuit)
        if row is None:
            row = ws.max_row + 1
            ws.cell(row=row, column=col_cuit).value = cuit

        for api_key, excel_col in API_TO_EXCEL.items():
            if api_key == "cuit":
                continue
            if api_key not in data:
                continue

            col = hmap[excel_col]
            val = data[api_key]

            if excel_col in ("Punto Venta", "Tipo Comprobante", "Numero Actividad"):
                ws.cell(row=row, column=col).value = _norm_int(val)
            elif excel_col == "¿Con detalle de Items?":
                ws.cell(row=row, column=col).value = _bool_to_si_no(val)
            else:
                ws.cell(row=row, column=col).value = "" if val is None else str(val)

        wb.save(excel_path)
        return _row_to_api(ws, row, hmap)


def sync_row2_all_tipos(
    excel_path: str,
    *,
    cuit: int,
    razon_social: str = "",
    domicilio_comercial: str = "",
    condicion_iva: str = "",
    #ingresos_brutos: Optional[str] = None,
    #fecha_inicio_actividades: Optional[str] = None,
    punto_venta: int,
    numero_actividad: int,
) -> Dict[str, Any]:
    
    lock_path = excel_path + ".lock"
    with _LockFile(lock_path):
        wb = _open_wb(excel_path)

        changed: Dict[str, Any] = {}
        for tipo, sheet in SHEET_BY_TIPO.items():
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            hmap = _header_map(ws)
            _ensure_headers(ws, hmap)

            r = 2  # <- tu requisito
            ws.cell(row=r, column=hmap["Cuit"]).value = int(cuit)
            ws.cell(row=r, column=hmap["Punto Venta"]).value = int(punto_venta)
            ws.cell(row=r, column=hmap["Numero Actividad"]).value = int(numero_actividad)

            ws.cell(row=r, column=hmap["Razón Social"]).value = razon_social or ""
            ws.cell(row=r, column=hmap["Domicilio Comercial"]).value = domicilio_comercial or ""
            ws.cell(row=r, column=hmap["Condición IVA"]).value = condicion_iva or ""
            # if ingresos_brutos is not None:
            #     ws.cell(row=r, column=hmap["Ingresos Brutos"]).value = _norm_text(ingresos_brutos)
            # if fecha_inicio_actividades is not None:
            #     ws.cell(row=r, column=hmap["Fecha de Inicio de Actividades"]).value = _norm_text(fecha_inicio_actividades)

            changed[tipo] = {"sheet": sheet, "row": r}

        wb.save(excel_path)
        return {"excel_path": excel_path, "changed": changed}
